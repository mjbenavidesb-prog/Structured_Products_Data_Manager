"""
Styled Excel report generator.
One sheet, grouped by the selected view dimension, matching the reference layout.
"""

from io import BytesIO
from datetime import date
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers as xl_numbers
    )
    from openpyxl.utils import get_column_letter
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


# ── Helpers ────────────────────────────────────────────────────────────────────

def _hex_fill(hex_color: str) -> "PatternFill":
    c = hex_color.lstrip("#")
    return PatternFill("solid", fgColor=c)


def _thin_border() -> "Border":
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _safe(v, default="—"):
    if v is None:
        return default
    try:
        if str(v).strip() in ("", "nan", "None", "NaN"):
            return default
    except Exception:
        pass
    return v


def _pct(v):
    if v is None:
        return "—"
    try:
        f = float(v)
        if abs(f) <= 1.5:
            f *= 100
        return f"{f:.2f}%"
    except (TypeError, ValueError):
        return "—"


def _num(v):
    if v is None:
        return "—"
    try:
        f = float(v)
        return f"{f:,.2f}" if f != f else "—"
    except (TypeError, ValueError):
        return "—"


def _usd(v):
    if v is None:
        return 0
    try:
        f = float(v)
        return 0 if f != f else f
    except (TypeError, ValueError):
        return 0


def _build_strikes(row: dict) -> str:
    parts = []
    for i in range(1, 5):
        u = _safe(row.get(f"underlying_{i}"), None)
        s = row.get(f"strike_{i}")
        if u and s not in (None, "", "nan", "None"):
            try:
                parts.append(f"{u}: {float(s):,.2f}")
            except (TypeError, ValueError):
                pass
    return " / ".join(parts) if parts else "—"


def _build_rendimiento(row: dict) -> tuple[str, str]:
    parts = []
    worst_val, worst_u = None, None
    for i in range(1, 5):
        u  = _safe(row.get(f"underlying_{i}"), None)
        s  = row.get(f"strike_{i}")
        sp = row.get(f"spot_{i}")
        if not u:
            continue
        try:
            s_f  = float(s)
            sp_f = float(sp)
            if s_f == 0:
                continue
            ret = (sp_f / s_f - 1) * 100
            parts.append(f"{u}: {ret:+.2f}%")
            if worst_val is None or ret < worst_val:
                worst_val, worst_u = ret, u
        except (TypeError, ValueError):
            pass
    rend   = " / ".join(parts) if parts else "—"
    worst  = f"{worst_u}: {worst_val:+.2f}%" if worst_val is not None else "—"
    return rend, worst


def _next_autocall(row: dict) -> str:
    today = date.today()
    future = []
    for i in range(1, 11):
        v = row.get(f"fecha_autocall_{i}")
        if v is None:
            continue
        try:
            d = pd.to_datetime(str(v), dayfirst=True, errors="coerce")
            if pd.notna(d) and d.date() > today:
                future.append(d.date())
        except Exception:
            pass
    if not future:
        return "—"
    nxt = min(future)
    return nxt.strftime("%d-%b-%y")


def _dominant_country(row: dict) -> str:
    mapping = {
        "Peru":     "monto_peru",
        "Chile":    "monto_chile",
        "Colombia": "monto_colombia",
        "USA":      "monto_usa",
    }
    best, best_v = "—", -1
    for country, col in mapping.items():
        v = _usd(row.get(col))
        if v > best_v:
            best, best_v = country, v
    return best


_GROUP_COL = {
    "AUM por Asset Class":      "asset_class",
    "AUM por Vehículo":         "vehiculo",
    "AUM por Estrategia":       "estrategia",        # "Estrategia" col from CSV
    "AUM por Perfil de riesgo": "perfil",
    "AUM por Contraparte":      "contraparte",
    "AUM por Segmento":         "tipo",
    "AUM por País":             "_dominant_country",
    "Portfolio Completo":       "estrategia",
}

_COLS = [
    ("Nombre del producto",  "nombre_producto"),
    ("Monto en USD",         "_monto"),
    ("Tipo",                 "tipo"),
    ("Moneda",               "moneda"),
    ("Contraparte",          "contraparte"),
    ("ISIN",                 "isin"),
    ("Fecha de Inicio",      "fecha_inicio"),
    ("Fecha de Obs. Final",  "fecha_obs_final"),
    ("Fecha de Vencimiento", "fecha_vencimiento"),
    ("Próximo Autocall",     "_next_autocall"),
    ("Cupón Anual",          "_cupon"),
    ("Ganancia Máxima",      "ganancia_maxima"),
    ("Barrera",              "_barrera"),
    ("Palanca",              "factor_participacion"),
    ("Autocall Trigger",     "_trigger"),
    ("Strikes",              "_strikes"),
    ("Rendimiento",          "_rendimiento"),
    ("Worst of",             "_worst"),
]

_COL_WIDTHS = [
    28, 14, 14, 8, 14, 18, 13, 13, 13, 13, 11, 12, 9, 9, 13, 36, 36, 18
]


# ── Main generator ─────────────────────────────────────────────────────────────

def generate_excel_report(
    df: pd.DataFrame,
    view: str,
    primary_hex: str = "#1F3864",
    secondary_hex: str = "#7F0000",
    filter_label: str = "",
) -> bytes:
    if not _OPENPYXL:
        raise ImportError("openpyxl is required for Excel export.")

    # ── Precompute derived columns ─────────────────────────────────────────────
    rows = df.to_dict("records")
    for r in rows:
        r["_monto"]        = _usd(r.get("monto_total"))
        r["_cupon"]        = _pct(r.get("cupon_contingente") or r.get("cupon_fijo"))
        bk = r.get("barrera_capital")
        if bk is not None:
            try:
                bk_f = float(bk)
                r["_barrera"] = f"{(1 - (bk_f if bk_f <= 1 else bk_f/100)) * 100:.0f}%"
            except (TypeError, ValueError):
                r["_barrera"] = "—"
        else:
            r["_barrera"] = "—"
        r["_trigger"]      = _pct(r.get("trigger_autocall"))
        r["_strikes"]      = _build_strikes(r)
        rend, worst        = _build_rendimiento(r)
        r["_rendimiento"]  = rend
        r["_worst"]        = worst
        r["_next_autocall"] = _next_autocall(r)
        r["_dominant_country"] = _dominant_country(r)

    # ── Group products ─────────────────────────────────────────────────────────
    group_col = _GROUP_COL.get(view, "tipo")
    # Fallback chain: estrategia → tipo_estructura → tipo
    if group_col in ("estrategia", "tipo_estructura"):
        has_data = any(
            r.get(group_col) and
            str(r.get(group_col)).strip() not in ("", "nan", "None")
            for r in rows
        )
        if not has_data and group_col == "estrategia":
            group_col = "tipo_estructura"
            has_data = any(
                r.get(group_col) and
                str(r.get(group_col)).strip() not in ("", "nan", "None")
                for r in rows
            )
        if not has_data:
            group_col = "tipo"

    groups: dict[str, list] = {}
    for r in rows:
        key = str(_safe(r.get(group_col), "Sin clasificar"))
        groups.setdefault(key, []).append(r)

    # ── Style setup ────────────────────────────────────────────────────────────
    pri   = primary_hex.lstrip("#")
    sec   = secondary_hex.lstrip("#")
    light = "F2F4F8"   # alternating row bg

    fill_title   = _hex_fill(pri)
    fill_section = _hex_fill(sec)
    fill_header  = _hex_fill(pri)
    fill_total   = PatternFill("solid", fgColor="D9D9D9")
    fill_alt     = PatternFill("solid", fgColor=light)
    fill_white   = PatternFill("solid", fgColor="FFFFFF")

    font_title   = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    font_info    = Font(name="Calibri", italic=True, color="595959", size=11)
    font_section = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    font_header  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    font_data    = Font(name="Calibri", size=11)
    font_total   = Font(name="Calibri", bold=True, size=11)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    align_right  = Alignment(horizontal="right",  vertical="center", wrap_text=True)
    border       = _thin_border()

    n_cols = len(_COLS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Set column widths
    for i, w in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    current_row = 1

    # ── Title row ──────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=n_cols)
    cell = ws.cell(current_row, 1,
                   f"Notas Estructuradas{' — ' + filter_label if filter_label else ''}")
    cell.font      = font_title
    cell.fill      = fill_title
    cell.alignment = align_center
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    # ── Info row ───────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=n_cols)
    cell = ws.cell(current_row, 1,
                   f"Información al {date.today().strftime('%d / %B / %y').title()}  ·  "
                   f"Vista: {view}  ·  Total: {len(rows)} productos")
    cell.font      = font_info
    cell.alignment = align_left
    ws.row_dimensions[current_row].height = 16
    current_row += 1

    # ── Groups ────────────────────────────────────────────────────────────────
    for group_name, group_rows in sorted(groups.items()):
        # Section header
        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=n_cols)
        cell = ws.cell(current_row, 1, group_name.upper())
        cell.font      = font_section
        cell.fill      = fill_section
        cell.alignment = align_left
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Column headers
        for col_idx, (header, _) in enumerate(_COLS, start=1):
            cell = ws.cell(current_row, col_idx, header)
            cell.font      = font_header
            cell.fill      = fill_header
            cell.alignment = align_center
            cell.border    = border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        group_total = 0.0

        # Data rows
        for i, r in enumerate(group_rows):
            fill = fill_alt if i % 2 == 1 else fill_white
            for col_idx, (_, field) in enumerate(_COLS, start=1):
                val = r.get(field, "—")
                if val is None or str(val).strip() in ("", "nan", "None", "NaN"):
                    val = "—"
                cell = ws.cell(current_row, col_idx, val)
                cell.font   = font_data
                cell.fill   = fill
                cell.border = border
                # Right-align numeric columns
                if col_idx == 2:   # Monto en USD
                    cell.alignment = align_right
                    if isinstance(val, (int, float)):
                        cell.number_format = '#,##0'
                elif col_idx in (11, 12, 13, 14, 15):  # pct columns
                    cell.alignment = align_right
                elif col_idx in (16, 17, 18):  # strikes / rendimiento / worst
                    cell.alignment = align_left
                else:
                    cell.alignment = align_left
            group_total += r["_monto"]
            ws.row_dimensions[current_row].height = 28
            current_row += 1

        # Group total row
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(current_row, col_idx,
                           f"Total {group_name}" if col_idx == 1 else
                           (group_total if col_idx == 2 else ""))
            cell.font   = font_total
            cell.fill   = fill_total
            cell.border = border
            if col_idx == 2:
                cell.alignment   = align_right
                cell.number_format = '#,##0'
            else:
                cell.alignment = align_left
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # ── Grand total ───────────────────────────────────────────────────────────
    grand_total = sum(r["_monto"] for r in rows)
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=n_cols)
    cell = ws.cell(current_row, 1,
                   f"TOTAL PORTAFOLIO: USD {grand_total:,.0f}")
    cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    cell.fill      = fill_title
    cell.alignment = align_right
    cell.border    = _thin_border()
    ws.row_dimensions[current_row].height = 20

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
