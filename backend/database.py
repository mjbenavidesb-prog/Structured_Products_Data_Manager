import sqlite3
import pandas as pd
from pathlib import Path
import json

DB_PATH = Path(__file__).parent.parent / "data" / "products.db"
CSV_PATH = Path(__file__).parent.parent / "data" / "Base Estructurados.csv"


def get_connection():
    DB_PATH.parent.mkdir(exist_ok=True)
    return sqlite3.connect(str(DB_PATH))


def init_db():
    conn = get_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre_producto TEXT,
            jurisdiccion TEXT,
            vehiculo TEXT,
            on_off TEXT,
            entidad TEXT,
            clasificacion_cliente TEXT,
            tipo_cliente TEXT,
            tipo TEXT,
            status TEXT,
            comentario TEXT,
            monto_cc_saf REAL,
            monto_credibolsa REAL,
            monto_agf REAL,
            monto_cc_colombia REAL,
            monto_asb_bank REAL,
            monto_asb_valores REAL,
            monto_ccc_llc REAL,
            monto_peru REAL,
            monto_chile REAL,
            monto_colombia REAL,
            monto_usa REAL,
            monto_total REAL,
            monto_bp_peru REAL,
            monto_bp_chile REAL,
            monto_bp_colombia REAL,
            monto_bp_us REAL,
            monto_ria REAL,
            monto_w9 REAL,
            monto_enalta REAL,
            monto_bex REAL,
            monto_consumo REAL,
            monto_juridicos REAL,
            monto_mfo REAL,
            monto_vicctus REAL,
            monto_otros REAL,
            monto_tyba REAL,
            fecha_ejecucion TEXT,
            fecha_strike TEXT,
            fecha_inicio TEXT,
            fecha_obs_final TEXT,
            fecha_vencimiento TEXT,
            fecha_pago_tentativo TEXT,
            fecha_pago_maximo TEXT,
            fecha_pago_cliente TEXT,
            dias_habiles_pago INTEGER,
            moneda TEXT,
            formato TEXT,
            contraparte TEXT,
            isin TEXT,
            isin_cavali TEXT,
            nemonico TEXT,
            producto_id TEXT,
            reoffer TEXT,
            fecha_emision TEXT,
            contraparte_derivado TEXT,
            notional_derivado REAL,
            prima_pct REAL,
            prima_usd REAL,
            formato_subyacente TEXT,
            underlying_1 TEXT,
            underlying_2 TEXT,
            underlying_3 TEXT,
            underlying_4 TEXT,
            strike_1 REAL,
            strike_2 REAL,
            strike_3 REAL,
            strike_4 REAL,
            spot_1 REAL,
            spot_2 REAL,
            spot_3 REAL,
            spot_4 REAL,
            peso_1 REAL,
            peso_2 REAL,
            peso_3 REAL,
            peso_4 REAL,
            rendimiento_1 REAL,
            rendimiento_2 REAL,
            rendimiento_3 REAL,
            rendimiento_4 REAL,
            rendimiento_total REAL,
            peor_subyacente TEXT,
            estrategia TEXT,
            tipo_estructura TEXT,
            asset_class TEXT,
            plazo_meses INTEGER,
            plazo_remanente_dias INTEGER,
            perfil TEXT,
            cupon_fijo REAL,
            cupon_contingente REAL,
            cap REAL,
            factor_participacion REAL,
            ganancia_maxima TEXT,
            trigger_autocall REAL,
            barrera_cupon REAL,
            barrera_capital REAL,
            tipo_caida TEXT,
            fecha_sin_autocall TEXT,
            fecha_autocall_1 TEXT,
            fecha_autocall_2 TEXT,
            fecha_autocall_3 TEXT,
            fecha_autocall_4 TEXT,
            fecha_autocall_5 TEXT,
            fecha_autocall_6 TEXT,
            fecha_autocall_7 TEXT,
            fecha_autocall_8 TEXT,
            fecha_autocall_9 TEXT,
            fecha_autocall_10 TEXT,
            fecha_obs_final_ac TEXT,
            proximo_autocall TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS config (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS aum_snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT,
            monto_total REAL,
            monto_peru REAL,
            monto_chile REAL,
            monto_colombia REAL,
            monto_usa REAL,
            n_productos INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    conn.commit()
    conn.close()


def _parse_amount(val, is_pct=False):
    if pd.isna(val) or val == "" or val == " ":
        return None
    try:
        s = str(val).strip()
        has_pct = s.endswith("%")
        clean = s.replace("%", "").replace(",", "").replace(" ", "").replace("(", "-").replace(")", "")
        f = float(clean)
        # Percentage columns: store as decimal (10.74% → 0.1074)
        if has_pct or is_pct:
            f = f / 100.0
        return f
    except Exception:
        return None


def _parse_pct(val):
    if pd.isna(val) or val == "":
        return None
    try:
        clean = str(val).replace("%", "").replace(",", ".").strip()
        return float(clean) / 100 if abs(float(clean)) > 1 else float(clean)
    except Exception:
        return None


_COL_MAP = {
    "Nombre del producto": "nombre_producto",
    "Jurisdicción": "jurisdiccion",
    "Vehiculo": "vehiculo",
    "ON/OFF": "on_off",
    "Entidad": "entidad",
    "Clasificacion de cliente": "clasificacion_cliente",
    "Tipo de Cliente": "tipo_cliente",
    "Tipo (Aut/Vcto)": "tipo",
    "Status": "status",
    "Comentario Adicional": "comentario",
    "Monto CC SAF": "monto_cc_saf",
    "Monto Credibolsa": "monto_credibolsa",
    "Monto AGF": "monto_agf",
    "Monto CC Colombia": "monto_cc_colombia",
    "Monto ASB Bank Corp": "monto_asb_bank",
    "Monto ASB Valores": "monto_asb_valores",
    "Monto Credicorp Capital LLC": "monto_ccc_llc",
    "Monto Peru": "monto_peru",
    "Monto Chile": "monto_chile",
    "Monto Colombia": "monto_colombia",
    "Monto USA": "monto_usa",
    "Monto total": "monto_total",
    "Monto BP Perú": "monto_bp_peru",
    "Monto BP Chile": "monto_bp_chile",
    "Monto BP Colombia": "monto_bp_colombia",
    "Monto BP US": "monto_bp_us",
    "Monto RIA": "monto_ria",
    "Monto W9": "monto_w9",
    "Monto Enalta": "monto_enalta",
    "Monto BEX": "monto_bex",
    "Monto Consumo": "monto_consumo",
    "Monto Juridicos": "monto_juridicos",
    "Monto MFO": "monto_mfo",
    "Monto Vicctus": "monto_vicctus",
    "Monto otros": "monto_otros",
    "TYBA": "monto_tyba",
    "Fecha de Ejecución": "fecha_ejecucion",
    "Fecha de Strike": "fecha_strike",
    "Fecha de Inicio": "fecha_inicio",
    "Fecha de Obs, Final": "fecha_obs_final",
    "Fecha de Vencimiento": "fecha_vencimiento",
    "Fecha de pago tentativo": "fecha_pago_tentativo",
    "Fecha de pago máximo": "fecha_pago_maximo",
    "Moneda": "moneda",
    "Formato": "formato",
    "Contraparte": "contraparte",
    "ISIN": "isin",
    "ISIN Cavali": "isin_cavali",
    "Nemónico Comercial": "nemonico",
    "ID": "producto_id",
    "Reoffer": "reoffer",
    "Fecha de Emisión (Issue Date)": "fecha_emision",
    "Contraparte Derivado": "contraparte_derivado",
    "Notional Derivado": "notional_derivado",
    "Prima %": "prima_pct",
    "Prima USD": "prima_usd",
    "Formato Subyacente": "formato_subyacente",
    "Underlying 1": "underlying_1",
    "Underlying 2": "underlying_2",
    "Underlying 3": "underlying_3",
    "Underlying 4": "underlying_4",
    "Strike 1": "strike_1",
    "Strike 2": "strike_2",
    "Strike 3": "strike_3",
    "Strike 4": "strike_4",
    "Spot 1": "spot_1",
    "Spot 2": "spot_2",
    "Spot 3": "spot_3",
    "Spot 4": "spot_4",
    "Peso 1": "peso_1",
    "Peso 2": "peso_2",
    "Peso 3": "peso_3",
    "Peso 4": "peso_4",
    "Rendimiento 1": "rendimiento_1",
    "Rendimiento 2": "rendimiento_2",
    "Rendimiento 3": "rendimiento_3",
    "Rendimiento 4": "rendimiento_4",
    "Rendimiento Total": "rendimiento_total",
    "Peor Subyacente": "peor_subyacente",
    "Estrategia": "estrategia",
    "Tipo de Estructura": "tipo_estructura",
    "Asset Class": "asset_class",
    "Plazo (meses)": "plazo_meses",
    "Plazo Remanente (días)": "plazo_remanente_dias",
    "Perfil": "perfil",
    "Cupón Anual Fijo": "cupon_fijo",
    "Cupón Anual Contingente": "cupon_contingente",
    "Cap": "cap",
    "Factor de Participación": "factor_participacion",
    "Ganancia Maxima": "ganancia_maxima",
    "Trigger Autocall": "trigger_autocall",
    "Barrera cupón": "barrera_cupon",
    "Barrera Capital": "barrera_capital",
    "Tipo de caída": "tipo_caida",
    "Fecha sin autocall": "fecha_sin_autocall",
    "Fecha 1": "fecha_autocall_1",
    "Fecha 2": "fecha_autocall_2",
    "Fecha 3": "fecha_autocall_3",
    "Fecha 4": "fecha_autocall_4",
    "Fecha 5": "fecha_autocall_5",
    "Fecha 6": "fecha_autocall_6",
    "Fecha 7": "fecha_autocall_7",
    "Fecha 8": "fecha_autocall_8",
    "Fecha 9": "fecha_autocall_9",
    "Fecha 10": "fecha_autocall_10",
    "Fecha obs final": "fecha_obs_final_ac",
    "Próximo Autocall": "proximo_autocall",
}

_TEMPLATE_EXAMPLE = {
    "Nombre del producto":          "Autocall Range Accrual Equity CCXXX",
    "Status":                       "VIGENTE",
    "Tipo (Aut/Vcto)":              "Autocall",
    "Moneda":                       "USD",
    "Monto total":                  "1000000",
    "Monto Peru":                   "1000000",
    "Contraparte":                  "Citigroup",
    "ISIN":                         "XS1234567890",
    "Asset Class":                  "Equity",
    "Estrategia":                   "Range Accrual",
    "Perfil":                       "Moderado",
    "Fecha de Strike":              "18/12/2024",
    "Fecha de Inicio":              "18/12/2024",
    "Fecha de Vencimiento":         "18/12/2026",
    "Plazo (meses)":                "24",
    "Underlying 1":                 "IWM",
    "Underlying 2":                 "XLF",
    "Underlying 3":                 "XLV",
    "Underlying 4":                 "XLP",
    "Strike 1":                     "203.21",
    "Strike 2":                     "42.50",
    "Strike 3":                     "138.00",
    "Strike 4":                     "75.20",
    "Spot 1":                       "220.50",
    "Spot 2":                       "48.30",
    "Spot 3":                       "145.00",
    "Spot 4":                       "80.10",
    "Cupón Anual Contingente":      "10.74%",
    "Trigger Autocall":             "100%",
    "Barrera cupón":                "75%",
    "Barrera Capital":              "75%",
    "Fecha 1":                      "18/03/2025",
    "Fecha 2":                      "18/06/2025",
    "Fecha 3":                      "18/09/2025",
    "Fecha 4":                      "18/12/2025",
    "Fecha 5":                      "18/03/2026",
    "Fecha 6":                      "18/06/2026",
}

_AMOUNT_COLS = {c for c in _COL_MAP if "Monto" in c or c in ("TYBA", "Notional Derivado", "Prima USD")}
_PCT_COLS = {
    "Prima %", "Cupón Anual Fijo", "Cupón Anual Contingente", "Cap",
    "Factor de Participación", "Trigger Autocall", "Barrera cupón", "Barrera Capital",
    "Peso 1", "Peso 2", "Peso 3", "Peso 4",
    "Rendimiento 1", "Rendimiento 2", "Rendimiento 3", "Rendimiento 4",
    "Rendimiento Total",
}
_PRICE_COLS = {
    "Strike 1", "Strike 2", "Strike 3", "Strike 4",
    "Spot 1",   "Spot 2",   "Spot 3",   "Spot 4",
    "Plazo (meses)", "Plazo Remanente (días)",
}


def _import_dataframe(df: pd.DataFrame, conn) -> int:
    """Parse a portfolio DataFrame and insert all non-empty rows into products. Returns count."""
    rows_inserted = 0
    for _, row in df.iterrows():
        nombre = row.get("Nombre del producto", "")
        if pd.isna(nombre) or str(nombre).strip() == "":
            continue

        record = {}
        for csv_col, db_col in _COL_MAP.items():
            val = row.get(csv_col, None)
            if csv_col in _AMOUNT_COLS:
                record[db_col] = _parse_amount(val, is_pct=False)
            elif csv_col in _PCT_COLS:
                record[db_col] = _parse_amount(val, is_pct=True)
            elif csv_col in _PRICE_COLS:
                record[db_col] = _parse_amount(val, is_pct=False)
            else:
                record[db_col] = str(val).strip() if not pd.isna(val) and val != "" else None

        record["dias_habiles_pago"] = None
        record["fecha_pago_cliente"] = None

        cols = ", ".join(record.keys())
        placeholders = ", ".join(["?"] * len(record))
        conn.execute(f"INSERT INTO products ({cols}) VALUES ({placeholders})", list(record.values()))
        rows_inserted += 1

    return rows_inserted


def _read_portfolio_df(buf, filename: str) -> pd.DataFrame:
    """
    Read a portfolio file (CSV or Excel) into a DataFrame.
    Tries skiprows=4 first (company format with 4 header rows), then skiprows=0.
    """
    fname = filename.lower()
    if fname.endswith(".xlsx") or fname.endswith(".xls"):
        df = pd.read_excel(buf, skiprows=4, header=0)
        if "Nombre del producto" not in df.columns:
            buf.seek(0) if hasattr(buf, "seek") else None
            df = pd.read_excel(buf, skiprows=0, header=0)
    else:
        raw = buf if isinstance(buf, bytes) else buf.read()
        for enc in ("utf-8-sig", "latin-1", "cp1252"):
            try:
                from io import BytesIO as _BIO
                df = pd.read_csv(_BIO(raw), sep=";", encoding=enc,
                                 skiprows=4, header=0, low_memory=False)
                if "Nombre del producto" not in df.columns:
                    df = pd.read_csv(_BIO(raw), sep=",", encoding=enc,
                                     skiprows=0, header=0, low_memory=False)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("No se pudo leer el archivo CSV con ninguna codificación.")
    return df


def seed_from_file(file_bytes: bytes, filename: str) -> int:
    """Import products from uploaded Excel/CSV bytes. Always inserts (no skip if DB has data)."""
    from io import BytesIO
    df = _read_portfolio_df(BytesIO(file_bytes), filename)
    conn = get_connection()
    result = _import_dataframe(df, conn)
    conn.commit()
    conn.close()
    return result


def reseed_from_file(file_bytes: bytes, filename: str) -> int:
    """Clear all products and reimport from uploaded Excel/CSV bytes."""
    conn = get_connection()
    conn.execute("DELETE FROM products")
    conn.commit()
    conn.close()
    return seed_from_file(file_bytes, filename)


def reseed_from_csv():
    """Force reseed from the on-disk CSV (kept for backward compatibility)."""
    conn = get_connection()
    conn.execute("DELETE FROM products")
    conn.commit()
    conn.close()
    return seed_from_csv()


def seed_from_csv():
    if not CSV_PATH.exists():
        return False
    conn = get_connection()
    count = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    if count > 0:
        conn.close()
        return True
    try:
        file_bytes = CSV_PATH.read_bytes()
        df = _read_portfolio_df(file_bytes, CSV_PATH.name)
    except Exception:
        return False
    result = _import_dataframe(df, conn)
    conn.commit()
    conn.close()
    return result


_NUMERIC_COLS = [
    "monto_total", "monto_peru", "monto_chile", "monto_colombia", "monto_usa",
    "monto_cc_saf", "monto_credibolsa", "monto_agf", "monto_cc_colombia",
    "monto_asb_bank", "monto_asb_valores", "monto_ccc_llc",
    "monto_bp_peru", "monto_bp_chile", "monto_bp_colombia", "monto_bp_us",
    "monto_ria", "monto_w9", "monto_enalta", "monto_bex",
    "monto_consumo", "monto_juridicos", "monto_mfo", "monto_vicctus",
    "monto_otros", "monto_tyba",
    "strike_1", "strike_2", "strike_3", "strike_4",
    "spot_1", "spot_2", "spot_3", "spot_4",
    "peso_1", "peso_2", "peso_3", "peso_4",
    "rendimiento_1", "rendimiento_2", "rendimiento_3", "rendimiento_4",
    "rendimiento_total", "cupon_fijo", "cupon_contingente", "cap",
    "factor_participacion", "trigger_autocall", "barrera_cupon", "barrera_capital",
    "prima_pct", "prima_usd", "notional_derivado", "plazo_meses", "plazo_remanente_dias",
]


def get_all_products(status_filter=None):
    conn = get_connection()
    query = "SELECT * FROM products"
    if status_filter:
        placeholders = ",".join(["?"] * len(status_filter))
        query += f" WHERE status IN ({placeholders})"
        df = pd.read_sql_query(query, conn, params=status_filter)
    else:
        df = pd.read_sql_query(query, conn)
    conn.close()
    for col in _NUMERIC_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def get_product_by_id(product_id):
    conn = get_connection()
    df = pd.read_sql_query("SELECT * FROM products WHERE id = ?", conn, params=[product_id])
    conn.close()
    return df.iloc[0] if not df.empty else None


def update_spots(updates: dict):
    """updates = {product_id: {spot_1: val, spot_2: val, ...}}"""
    conn = get_connection()
    for pid, fields in updates.items():
        set_clause = ", ".join([f"{k} = ?" for k in fields])
        values = list(fields.values()) + [pid]
        conn.execute(f"UPDATE products SET {set_clause}, updated_at = CURRENT_TIMESTAMP WHERE id = ?", values)
    conn.commit()
    conn.close()


def insert_product(record: dict):
    conn = get_connection()
    cols = ", ".join(record.keys())
    placeholders = ", ".join(["?"] * len(record))
    cursor = conn.execute(f"INSERT INTO products ({cols}) VALUES ({placeholders})", list(record.values()))
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()
    return new_id


def update_product(product_id: int, fields: dict):
    conn = get_connection()
    set_clause = ", ".join([f"{k} = ?" for k in fields])
    values = list(fields.values()) + [product_id]
    conn.execute(f"UPDATE products SET {set_clause}, updated_at = CURRENT_TIMESTAMP WHERE id = ?", values)
    conn.commit()
    conn.close()


def get_config(key: str, default=None):
    conn = get_connection()
    row = conn.execute("SELECT value FROM config WHERE key = ?", [key]).fetchone()
    conn.close()
    if row:
        try:
            return json.loads(row[0])
        except Exception:
            return row[0]
    return default


def set_config(key: str, value):
    conn = get_connection()
    conn.execute("INSERT OR REPLACE INTO config (key, value) VALUES (?, ?)", [key, json.dumps(value)])
    conn.commit()
    conn.close()


def save_aum_snapshot():
    df = get_all_products(status_filter=["VIGENTE", "POR EJECUTAR", "AUTOCALL"])
    if df.empty:
        return
    from datetime import date
    conn = get_connection()
    conn.execute("""
        INSERT INTO aum_snapshots (fecha, monto_total, monto_peru, monto_chile, monto_colombia, monto_usa, n_productos)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, [
        str(date.today()),
        df["monto_total"].sum(),
        df["monto_peru"].sum(),
        df["monto_chile"].sum(),
        df["monto_colombia"].sum(),
        df["monto_usa"].sum(),
        len(df),
    ])
    conn.commit()
    conn.close()


def get_aum_history():
    conn = get_connection()
    df = pd.read_sql_query("SELECT * FROM aum_snapshots ORDER BY fecha", conn)
    conn.close()
    return df


_TEMPLATE_COLS = [
    # Identificación
    "Nombre del producto", "Status", "Tipo (Aut/Vcto)", "Moneda", "Contraparte",
    "ISIN", "Asset Class", "Estrategia", "Tipo de Estructura", "Perfil",
    # Fechas
    "Fecha de Strike", "Fecha de Inicio", "Fecha de Obs, Final", "Fecha de Vencimiento",
    "Plazo (meses)",
    # Montos (universales)
    "Monto total", "Monto Peru", "Monto Chile", "Monto Colombia", "Monto USA",
    # Subyacentes
    "Underlying 1", "Underlying 2", "Underlying 3", "Underlying 4",
    "Strike 1", "Strike 2", "Strike 3", "Strike 4",
    "Spot 1",   "Spot 2",   "Spot 3",   "Spot 4",
    "Peso 1",   "Peso 2",   "Peso 3",   "Peso 4",
    # Estructura financiera
    "Cupón Anual Fijo", "Cupón Anual Contingente", "Cap",
    "Factor de Participación", "Ganancia Maxima",
    "Trigger Autocall", "Barrera cupón", "Barrera Capital",
    # Rendimientos
    "Rendimiento 1", "Rendimiento 2", "Rendimiento 3", "Rendimiento 4",
    "Rendimiento Total", "Peor Subyacente",
    # Fechas de observación autocall
    "Fecha 1", "Fecha 2", "Fecha 3", "Fecha 4", "Fecha 5",
    "Fecha 6", "Fecha 7", "Fecha 8", "Fecha 9", "Fecha 10",
]


def generate_template_xlsx() -> bytes:
    """
    Return bytes of an Excel template with all expected column headers and one
    example row. Users download this, fill in their data, and re-upload it.
    """
    from io import BytesIO
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        # ── Valid values for dropdowns ─────────────────────────────────────────
        _DROPDOWNS = {
            "Status":           ["VIGENTE", "VENCIDO", "AUTOCALL", "POR EJECUTAR"],
            "Tipo (Aut/Vcto)":  ["Autocall", "Vencimiento"],
            "Moneda":           ["USD", "PEN", "EUR", "CLP", "COP"],
            "Asset Class":      ["Equity", "Fixed Income", "Multi-Asset",
                                 "Commodities", "FX"],
            "Perfil":           ["Conservador", "Moderado", "Agresivo",
                                 "Muy Agresivo"],
            "Estrategia":       ["Range Accrual", "Phoenix", "Barrier Reverse Conv.",
                                 "Participation", "Capital Protegido", "Digital"],
            "Tipo de Estructura": ["Worst of", "Single Asset", "Basket"],
        }
        # Subyacentes: lista larga → hoja de referencia
        _UNDERLYINGS = [
            "IWM", "XLF", "XLV", "XLP", "XLE", "XLK", "XLI", "XLB",
            "SPY", "QQQ", "DIA", "EEM", "GLD", "TLT", "HYG",
            "^GSPC", "^NDX", "^RUT", "SX5E", "NKY", "HSI",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Portafolio"

        # ── Hidden reference sheet for long dropdown lists ─────────────────────
        ref_ws = wb.create_sheet("_Ref")
        ref_ws.sheet_state = "hidden"
        for ri, u in enumerate(_UNDERLYINGS, 1):
            ref_ws.cell(ri, 1, u)
        ref_range = f"_Ref!$A$1:$A${len(_UNDERLYINGS)}"

        headers = _TEMPLATE_COLS

        hdr_fill    = PatternFill("solid", fgColor="1F3864")
        hdr_font    = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
        ex_font     = Font(name="Calibri", size=9, italic=True, color="595959")
        thin_side   = Side(style="thin", color="CCCCCC")
        thin_border = Border(left=thin_side, right=thin_side,
                             top=thin_side, bottom=thin_side)
        center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        # ── Header row ─────────────────────────────────────────────────────────
        col_index = {}  # column name → letter
        for ci, col in enumerate(headers, 1):
            cell = ws.cell(1, ci, col)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = center
            cell.border    = thin_border
            ltr = get_column_letter(ci)
            col_index[col] = ltr
            ws.column_dimensions[ltr].width = max(12, min(28, len(col) + 3))
        ws.row_dimensions[1].height = 32

        # ── Example row ────────────────────────────────────────────────────────
        for ci, col in enumerate(headers, 1):
            val = _TEMPLATE_EXAMPLE.get(col, "")
            cell = ws.cell(2, ci, val)
            cell.font      = ex_font
            cell.alignment = left
            cell.border    = thin_border
        ws.row_dimensions[2].height = 18

        # ── Dropdowns for categorical columns ──────────────────────────────────
        for col_name, options in _DROPDOWNS.items():
            if col_name not in col_index:
                continue
            ltr = col_index[col_name]
            formula = '"' + ",".join(options) + '"'
            dv = DataValidation(type="list", formula1=formula,
                                allow_blank=True, showDropDown=False)
            dv.error      = f"Valor no válido. Opciones: {', '.join(options)}"
            dv.errorTitle = "Valor inválido"
            dv.prompt     = f"Selecciona: {', '.join(options)}"
            dv.add(f"{ltr}2:{ltr}500")
            ws.add_data_validation(dv)

        # ── Dropdowns for underlying columns (from hidden ref sheet) ───────────
        for und_col in ["Underlying 1", "Underlying 2",
                        "Underlying 3", "Underlying 4"]:
            if und_col not in col_index:
                continue
            ltr = col_index[und_col]
            dv = DataValidation(type="list", formula1=ref_range,
                                allow_blank=True, showDropDown=False)
            dv.prompt = "Ticker del subyacente (ej: IWM, XLF, SPY)"
            dv.add(f"{ltr}2:{ltr}500")
            ws.add_data_validation(dv)

        ws.freeze_panes = "A2"

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    except ImportError:
        # Fallback to pandas if openpyxl somehow not available
        buf = BytesIO()
        pd.DataFrame([_TEMPLATE_EXAMPLE], columns=_TEMPLATE_COLS).to_excel(
            buf, index=False
        )
        buf.seek(0)
        return buf.getvalue()
